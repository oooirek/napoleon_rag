import re
from datetime import date, datetime

import pdfplumber
from langchain_community.document_loaders import PyPDFLoader, UnstructuredWordDocumentLoader
from langchain_core.documents import Document
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

from src.rag.loaders.base import BaseLoader


def _clean_text(text: str) -> str:
	return " ".join(text.replace("\n", " ").split())


_HEADER_NOISE_RE = re.compile(r"^Столбец\d+$", re.IGNORECASE)
_TECH_ID_RE = re.compile(r"^[a-z][a-z0-9_]*$")
_ANNOTATION_RE = re.compile(r"^\(.+\)$")
# аннотации в шапке, которые не несут смысла и должны выкидываться из под-имени
_NOISE_ANNOTATION_RE = re.compile(
	r"\s*\((?:обязательное поле|не\s*обязательное поле|опционально)\)\s*$",
	re.IGNORECASE,
)
# горизонтальный merge шире этого числа колонок считаем «общим заголовком/инструкцией»,
# а не названием конкретной колонки
_WIDE_MERGE_COLS = 5
# при наследовании группового имени берём ближайший якорь не дальше этого расстояния
_GROUP_INHERIT_MAX_DIST = 1
_LABEL_MAP = {
	"п": "план", "план": "план", "plan": "план", "p": "план",
	"ф": "факт", "факт": "факт", "fact": "факт", "f": "факт",
}


def _is_blank(v) -> bool:
	if v is None:
		return True
	if isinstance(v, str) and not v.strip():
		return True
	return False


def _is_numeric_or_date(v) -> bool:
	if isinstance(v, bool):
		return False
	return isinstance(v, (int, float, datetime, date))


def _format_value(v) -> str:
	if isinstance(v, datetime):
		if v.hour == 0 and v.minute == 0 and v.second == 0 and v.microsecond == 0:
			return v.strftime("%d.%m.%Y")
		return v.strftime("%d.%m.%Y %H:%M")
	if isinstance(v, date):
		return v.strftime("%d.%m.%Y")
	if isinstance(v, float):
		if v.is_integer():
			return str(int(v))
		return f"{v:.6f}".rstrip("0").rstrip(".")
	return _clean_text(str(v))


def _normalize_label(s: str) -> str:
	return _LABEL_MAP.get(s.strip().lower(), s.strip().lower())


class PdfLoader(BaseLoader):
	def load(self) -> list[Document]:
		documents: list[Document] = []

		# основной текст берем старым проверенным способом
		text_docs = PyPDFLoader(file_path=self.file_path).load()
		for doc in text_docs:
			doc.metadata["type"] = "text"
		documents.extend(text_docs)

		# таблицы извлекаем отдельно (pdfplumber)
		table_idx = 0
		with pdfplumber.open(self.file_path) as pdf:
			for page_num, page in enumerate(pdf.pages):
				tables = page.extract_tables() or []
				for table in tables:
					if not table or not any(row for row in table):
						continue
					rows = []
					for row in table:
						if row and any(cell and str(cell).strip() for cell in row):
							cleaned_row = [_clean_text(str(cell)) if cell else "" for cell in row]
							rows.append(" | ".join(cleaned_row))
					if not rows:
						continue

					table_text = "Таблица:\n" + "\n".join(rows)
					metadata = {
						"source": self.file_path,
						"type": "table",
						"page": page_num + 1,
						"table_index": table_idx,
					}
					documents.append(Document(page_content=table_text, metadata=metadata))
					table_idx += 1

		return documents


# пока не работает
class WordLoader(BaseLoader):
	def load(self) -> list[Document]:
		return UnstructuredWordDocumentLoader(file_path=self.file_path).load()


def _looks_like_object_name(s: str) -> bool:
	return len(s) > 5 and any(ch.isalpha() for ch in s)


def _classify_header_row(values: list[str]) -> str:
	if not values:
		return "empty"
	if any(v.startswith("*") for v in values):
		return "instruction"
	if all(_HEADER_NOISE_RE.match(v) for v in values):
		return "noise"
	tech_cnt = sum(1 for v in values if _TECH_ID_RE.match(v))
	if tech_cnt >= max(2, len(values) // 2):
		return "tech"
	if len(values) <= 3 and all(len(v) > 30 for v in values):
		return "instruction"
	return "meaningful"


class ExcelLoader(BaseLoader):
	def load(self) -> list[Document]:
		wb = load_workbook(self.file_path, data_only=True)
		sheet_name = wb.sheetnames[0]
		ws = wb[sheet_name]

		n_rows = ws.max_row or 0
		n_cols = ws.max_column or 0
		if n_rows < 1 or n_cols < 1:
			return []

		# 1-indexed матрица значений
		cells: list[list] = [[None] * (n_cols + 1) for _ in range(n_rows + 1)]
		for row in ws.iter_rows(min_row=1, max_row=n_rows, max_col=n_cols):
			for cell in row:
				cells[cell.row][cell.column] = cell.value

		# разворачиваем merged-ячейки — значение из верхней-левой по всему диапазону.
		# дополнительно запоминаем, попадает ли ячейка под вертикальный merge внутри шапки
		# (тогда r_sub дублирует r_human и под-имя не нужно).
		merged_ranges = [range_boundaries(str(mr)) for mr in ws.merged_cells.ranges]
		vmerge_in_header: set[tuple[int, int]] = set()
		for min_col, min_row, max_col, max_row in merged_ranges:
			v = cells[min_row][min_col]
			for r in range(min_row, max_row + 1):
				for c in range(min_col, max_col + 1):
					cells[r][c] = v
			if max_row > min_row and max_col == min_col:
				for r in range(min_row + 1, max_row + 1):
					vmerge_in_header.add((r, min_col))

		# граница шапки — первая строка с >=2 numeric/date значениями
		header_end = 0
		for r in range(1, n_rows + 1):
			data_count = sum(1 for c in range(1, n_cols + 1) if _is_numeric_or_date(cells[r][c]))
			if data_count >= 2:
				header_end = r - 1
				break
		if header_end < 1:
			header_end = 1
		data_start = header_end + 1
		if data_start > n_rows:
			return []

		# классифицируем строки шапки
		row_cats: dict[int, str] = {}
		for r in range(1, header_end + 1):
			str_vals = [
				_clean_text(str(cells[r][c]))
				for c in range(1, n_cols + 1)
				if not _is_blank(cells[r][c])
			]
			str_vals = [v for v in str_vals if v]
			row_cats[r] = _classify_header_row(str_vals)

		# собираем технические ID (берём из самой нижней tech-строки)
		tech_ids: dict[int, str] = {}
		for r in range(1, header_end + 1):
			if row_cats[r] != "tech":
				continue
			for c in range(1, n_cols + 1):
				v = cells[r][c]
				if isinstance(v, str) and _TECH_ID_RE.match(v.strip()):
					tech_ids[c] = v.strip()

		# первые две meaningful строки → r_human (групповое имя) и r_sub (под-имя)
		meaningful_rows = [r for r in range(1, header_end + 1) if row_cats[r] == "meaningful"]
		r_human = meaningful_rows[0] if meaningful_rows else None
		r_sub = meaningful_rows[1] if len(meaningful_rows) > 1 else None

		human: list[str | None] = [None] * (n_cols + 2)
		sub: list[str | None] = [None] * (n_cols + 2)

		# r_human с приклейкой «оторванных» фрагментов (типа «Коор» + «динаты» → «Координаты»)
		if r_human is not None:
			prev_c: int | None = None
			for c in range(1, n_cols + 1):
				v = cells[r_human][c]
				if _is_blank(v):
					continue
				s = _clean_text(str(v))
				if not s or _HEADER_NOISE_RE.match(s) or _TECH_ID_RE.match(s):
					continue
				if (
					prev_c is not None
					and human[prev_c]
					and s[0].isalpha()
					and s[0].islower()
				):
					human[prev_c] = f"{human[prev_c]}{s}"
				else:
					human[c] = s
					prev_c = c

		# r_sub: чистим аннотации-шумы, выкидываем дубли r_human (вертикальный merge)
		if r_sub is not None:
			for c in range(1, n_cols + 1):
				v = cells[r_sub][c]
				if _is_blank(v):
					continue
				if (r_sub, c) in vmerge_in_header:
					continue
				s = _clean_text(str(v))
				if not s or _HEADER_NOISE_RE.match(s) or _TECH_ID_RE.match(s):
					continue
				s = _NOISE_ANNOTATION_RE.sub("", s).strip()
				if not s or _ANNOTATION_RE.match(s):
					continue
				if human[c] and (s == human[c] or s in human[c]):
					continue
				sub[c] = s

		# наследование группы: в пределах непрерывной «цепочки» колонок с под-именем
		# каждая колонка без r_human наследует ближайший якорь СЛЕВА; если слева в цепочке
		# якоря нет — берём ближайший справа
		group: list[str | None] = list(human)
		c = 1
		while c <= n_cols:
			if not sub[c]:
				c += 1
				continue
			start = c
			while c <= n_cols and sub[c]:
				c += 1
			end = c - 1
			anchors_in_chain = [k for k in range(start, end + 1) if human[k] and sub[k]]
			if not anchors_in_chain:
				continue
			for k in range(start, end + 1):
				if group[k]:
					continue
				left = [a for a in anchors_in_chain if a < k]
				right = [a for a in anchors_in_chain if a > k]
				if left and k - left[-1] <= _GROUP_INHERIT_MAX_DIST:
					group[k] = human[left[-1]]
				elif right and right[0] - k <= _GROUP_INHERIT_MAX_DIST:
					group[k] = human[right[0]]

		# итоговое имя колонки
		col_names: dict[int, str] = {}
		for c in range(1, n_cols + 1):
			g = group[c]
			s = sub[c]
			if g and s:
				col_names[c] = f"{g} - {s}"
			elif g:
				col_names[c] = g
			elif s:
				col_names[c] = s
			elif c in tech_ids:
				col_names[c] = tech_ids[c]

		# ключевая колонка: full_name из технических ID; иначе первая именованная с данными
		key_col: int | None = None
		for c, tid in tech_ids.items():
			if tid == "full_name":
				key_col = c
				break
		if key_col is None:
			for c in range(1, n_cols + 1):
				if c in col_names and not _is_blank(cells[data_start][c]):
					key_col = c
					break

		# вертикальные merged-ranges, охватывающие ключевую колонку — это явные группы
		key_vmerges: dict[int, int] = {}
		if key_col is not None:
			for min_col, min_row, max_col, max_row in merged_ranges:
				if not (min_col <= key_col <= max_col and max_row > min_row and min_row >= data_start):
					continue
				key_vmerges[min_row] = max(key_vmerges.get(min_row, min_row), max_row)

		# группы: если строка входит в merged range по ключевой колонке — группа = весь диапазон,
		# иначе каждая строка отдельный объект
		groups: list[tuple[int, int]] = []
		r = data_start
		while r <= n_rows:
			if all(_is_blank(cells[r][c]) for c in range(1, n_cols + 1)):
				r += 1
				continue
			end = key_vmerges.get(r, r)
			groups.append((r, end))
			r = end + 1

		documents: list[Document] = []
		for row_start, row_end in groups:
			rows = list(range(row_start, row_end + 1))

			# для групп из 2+ строк ищем колонку-разделитель (короткие уникальные маркеры)
			divider_col: int | None = None
			divider_labels: dict[int, str] = {}
			if len(rows) > 1:
				for c in range(1, n_cols + 1):
					if c == key_col:
						continue
					vals: list[tuple[int, str]] = []
					ok = True
					for rr in rows:
						v = cells[rr][c]
						if _is_blank(v):
							ok = False
							break
						s = _clean_text(str(v))
						if not s or len(s) > 8:
							ok = False
							break
						vals.append((rr, s))
					if not ok:
						continue
					if len({s for _, s in vals}) != len(vals):
						continue
					divider_col = c
					divider_labels = {rr: _normalize_label(s) for rr, s in vals}
					break

			# если ключевая колонка содержит «имя объекта» (строка с буквами),
			# выносим её в префикс и убираем из общего перечня
			obj_name: str | None = None
			if key_col is not None and not _is_blank(cells[row_start][key_col]):
				v = cells[row_start][key_col]
				formatted = _format_value(v)
				if _looks_like_object_name(formatted):
					obj_name = formatted

			parts: list[str] = []
			if obj_name:
				stripped = obj_name.rstrip(".")
				parts.append(f"Объект {stripped}.")

			for c in range(1, n_cols + 1):
				if c == divider_col:
					continue
				if c == key_col and obj_name:
					continue
				name = col_names.get(c)
				if not name:
					continue
				row_vals: dict[int, str] = {}
				for rr in rows:
					v = cells[rr][c]
					if _is_blank(v):
						continue
					row_vals[rr] = _format_value(v)
				if not row_vals:
					continue

				if len(row_vals) == 1 or len(set(row_vals.values())) == 1:
					parts.append(f"{name}: {next(iter(row_vals.values()))}.")
					continue

				for rr in rows:
					if rr not in row_vals:
						continue
					label = divider_labels.get(rr)
					if label:
						parts.append(f"{name} ({label}): {row_vals[rr]}.")
					else:
						parts.append(f"{name}: {row_vals[rr]}.")

			if not parts:
				continue

			documents.append(Document(
				page_content=" ".join(parts),
				metadata={
					"source": self.file_path,
					"type": "table",
					"sheet": sheet_name,
					"row_start": row_start,
					"row_end": row_end,
				},
			))

		return documents


# оставляем алиас для обратной совместимости
TableLoader = ExcelLoader
